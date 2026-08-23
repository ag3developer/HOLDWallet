"""
🗂️ HOLD Wallet - Account Export Service
========================================

Serviço para exportar dados completos da conta do usuário em múltiplos formatos:
- PDF com relatório formatado
- Excel com planilhas detalhadas
- JSON com dados brutos

Author: HOLD Wallet Team
"""

import io
import json
import logging
from datetime import datetime, timezone
from typing import Dict, Any, Optional, List, Tuple
from decimal import Decimal

from sqlalchemy.orm import Session
from sqlalchemy import func, desc

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Models
from app.models.user import User
from app.models.wallet import Wallet
from app.models.balance import WalletBalance
from app.models.instant_trade import InstantTrade, TradeStatus, PaymentMethod
from app.models.p2p import P2POrder, P2PMatch
from app.models.kyc import KYCVerification, KYCStatus
from app.models.wolkpay import WolkPayInvoice, WolkPayPayment

logger = logging.getLogger(__name__)


class DecimalEncoder(json.JSONEncoder):
    """JSON Encoder para Decimal"""
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        if isinstance(obj, datetime):
            return obj.isoformat()
        return super().default(obj)


class AccountExportService:
    """Serviço de exportação de dados da conta"""

    @staticmethod
    def collect_user_data(user: User, db: Session) -> Dict[str, Any]:
        """Coleta todos os dados do usuário"""
        logger.info(f"📊 Coletando dados da conta para {user.email}")

        # 1. Perfil do usuário
        user_profile = {
            "username": user.username,
            "email": user.email,
            "account_created": user.created_at.isoformat(),
            "account_status": "active" if user.is_active else "inactive",
            "email_verified": user.is_email_verified,
            "last_login": user.last_login.isoformat() if user.last_login else None,
            "is_admin": user.is_admin,
            "two_fa_enabled": bool(
                db.query(func.count()).select_from(
                    __import__('app.models.two_factor', fromlist=['TwoFactorAuth']).TwoFactorAuth
                ).filter(
                    __import__('app.models.two_factor', fromlist=['TwoFactorAuth']).TwoFactorAuth.user_id == user.id
                ).scalar()
            )
        }

        # 2. Wallets e Saldos
        wallets = db.query(Wallet).filter(Wallet.user_id == user.id).all()
        wallets_data = []
        total_balance_usd = Decimal('0')

        for wallet in wallets:
            balance = db.query(WalletBalance).filter(
                WalletBalance.wallet_id == wallet.id
            ).first()

            wallet_info = {
                "wallet_id": str(wallet.id),
                "address": wallet.address,
                "crypto": wallet.crypto,
                "network": wallet.network,
                "balance": float(balance.balance) if balance else 0,
                "balance_in_brl": float(balance.balance_in_brl) if balance else 0,
                "created_at": wallet.created_at.isoformat(),
                "is_active": wallet.is_active,
            }
            wallets_data.append(wallet_info)
            if balance:
                total_balance_usd += balance.balance

        # 3. Histórico de Transações OTC
        trades = db.query(InstantTrade).filter(
            (InstantTrade.buyer_id == user.id) | (InstantTrade.seller_id == user.id)
        ).order_by(desc(InstantTrade.created_at)).all()

        trades_data = []
        for trade in trades:
            trade_info = {
                "trade_id": str(trade.id),
                "type": "buy" if trade.buyer_id == user.id else "sell",
                "status": trade.status.value,
                "amount": float(trade.total_amount),
                "crypto": trade.crypto,
                "price_per_unit": float(trade.price_per_unit) if trade.price_per_unit else None,
                "quantity": float(trade.quantity),
                "payment_method": trade.payment_method.value if trade.payment_method else None,
                "counterparty": "buyer" if trade.buyer_id == user.id else "seller",
                "created_at": trade.created_at.isoformat(),
                "completed_at": trade.completed_at.isoformat() if trade.completed_at else None,
                "fee_amount": float(trade.fee_amount) if trade.fee_amount else 0,
            }
            trades_data.append(trade_info)

        # 4. Histórico P2P
        p2p_orders = db.query(P2POrder).filter(
            (P2POrder.creator_id == user.id) | (P2POrder.matched_with_id == user.id)
        ).order_by(desc(P2POrder.created_at)).all()

        p2p_data = []
        for order in p2p_orders:
            order_info = {
                "order_id": str(order.id),
                "order_type": order.order_type.value,
                "status": order.status.value if hasattr(order.status, 'value') else str(order.status),
                "amount": float(order.amount),
                "price": float(order.price),
                "total_value": float(order.amount * order.price),
                "crypto": order.crypto,
                "created_at": order.created_at.isoformat(),
                "completed_at": order.completed_at.isoformat() if order.completed_at else None,
            }
            p2p_data.append(order_info)

        # 5. KYC Verification
        kyc = db.query(KYCVerification).filter(
            KYCVerification.user_id == user.id
        ).first()

        kyc_data = None
        if kyc:
            kyc_data = {
                "status": kyc.status.value if hasattr(kyc.status, 'value') else str(kyc.status),
                "document_type": kyc.document_type,
                "verified_at": kyc.verified_at.isoformat() if kyc.verified_at else None,
                "expires_at": kyc.expires_at.isoformat() if kyc.expires_at else None,
                "daily_limit": float(kyc.daily_limit) if kyc.daily_limit else None,
                "monthly_limit": float(kyc.monthly_limit) if kyc.monthly_limit else None,
            }

        # 6. WolkPay Invoices
        invoices = db.query(WolkPayInvoice).filter(
            WolkPayInvoice.beneficiary_id == user.id
        ).order_by(desc(WolkPayInvoice.created_at)).all()

        invoices_data = []
        for invoice in invoices:
            invoice_info = {
                "invoice_id": str(invoice.id),
                "amount": float(invoice.amount),
                "status": invoice.status,
                "description": invoice.description,
                "due_date": invoice.due_date.isoformat() if invoice.due_date else None,
                "paid_at": invoice.paid_at.isoformat() if invoice.paid_at else None,
                "created_at": invoice.created_at.isoformat(),
            }
            invoices_data.append(invoice_info)

        # Compilar tudo
        export_data = {
            "export_date": datetime.now(timezone.utc).isoformat(),
            "export_version": "1.0",
            "user_profile": user_profile,
            "wallets": {
                "total_balance_usd": float(total_balance_usd),
                "count": len(wallets),
                "wallets": wallets_data,
            },
            "trades": {
                "count": len(trades),
                "total_volume_brl": float(sum(Decimal(t.get("amount", 0)) for t in trades_data)),
                "trades": trades_data,
            },
            "p2p": {
                "count": len(p2p_orders),
                "orders": p2p_data,
            },
            "kyc": kyc_data,
            "invoices": {
                "count": len(invoices),
                "invoices": invoices_data,
            },
        }

        logger.info(f"✅ Dados coletados: {len(trades)} trades, {len(wallets)} wallets, {len(p2p_orders)} P2P")
        return export_data

    @staticmethod
    def export_to_json(export_data: Dict[str, Any]) -> bytes:
        """Exporta dados para JSON"""
        logger.info("📄 Gerando JSON...")
        
        json_data = json.dumps(export_data, indent=2, cls=DecimalEncoder)
        return json_data.encode('utf-8')

    @staticmethod
    def export_to_pdf(user: User, export_data: Dict[str, Any]) -> bytes:
        """Exporta dados para PDF formatado"""
        if not REPORTLAB_AVAILABLE:
            logger.error("❌ ReportLab não instalado")
            raise ImportError("ReportLab não instalado. Execute: pip install reportlab")

        logger.info("📄 Gerando PDF...")

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30,
        )

        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=TA_CENTER,
        )
        title = Paragraph(f"Relatório de Exportação de Dados - {user.username}", title_style)
        elements.append(title)

        # Data de exportação
        export_date = export_data.get('export_date', '')
        info_text = f"<b>Data de Exportação:</b> {export_date}<br/><b>Email:</b> {user.email}"
        elements.append(Paragraph(info_text, styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))

        # Seção 1: Perfil do Usuário
        elements.append(Paragraph("<b>1. PERFIL DA CONTA</b>", styles['Heading2']))
        profile = export_data['user_profile']
        profile_data = [
            ['Username', profile['username']],
            ['Email', profile['email']],
            ['Criado em', profile['account_created']],
            ['Status', profile['account_status']],
            ['Email Verificado', 'Sim' if profile['email_verified'] else 'Não'],
            ['2FA Ativado', 'Sim' if profile['two_fa_enabled'] else 'Não'],
        ]
        profile_table = Table(profile_data, colWidths=[2*inch, 3*inch])
        profile_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f0f0')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        elements.append(profile_table)
        elements.append(Spacer(1, 0.3*inch))

        # Seção 2: Wallets
        elements.append(Paragraph("<b>2. CARTEIRAS</b>", styles['Heading2']))
        wallets_info = export_data['wallets']
        wallet_summary = f"Total de Carteiras: {wallets_info['count']} | Saldo Total (USD): ${wallets_info['total_balance_usd']:.2f}"
        elements.append(Paragraph(wallet_summary, styles['Normal']))
        
        if wallets_info['wallets']:
            wallet_rows = [['Crypto', 'Saldo', 'Endereço', 'Criada em']]
            for w in wallets_info['wallets']:
                wallet_rows.append([
                    w['crypto'],
                    f"${w['balance']:.4f}",
                    w['address'][:20] + '...',
                    w['created_at'][:10],
                ])
            wallet_table = Table(wallet_rows, colWidths=[1*inch, 1.2*inch, 1.5*inch, 1*inch])
            wallet_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(wallet_table)
        
        elements.append(Spacer(1, 0.2*inch))

        # Seção 3: Trades
        trades_info = export_data['trades']
        elements.append(Paragraph("<b>3. HISTÓRICO DE TRADES (OTC)</b>", styles['Heading2']))
        trades_summary = f"Total de Trades: {trades_info['count']} | Volume Total (BRL): R$ {trades_info['total_volume_brl']:.2f}"
        elements.append(Paragraph(trades_summary, styles['Normal']))
        
        if trades_info['trades']:
            trade_rows = [['Tipo', 'Status', 'Quantidade', 'Valor', 'Data']]
            for t in trades_info['trades'][:10]:  # Mostrar últimos 10
                trade_rows.append([
                    t['type'].upper(),
                    t['status'],
                    f"{t['quantity']:.4f} {t['crypto']}",
                    f"R$ {t['amount']:.2f}",
                    t['created_at'][:10],
                ])
            trade_table = Table(trade_rows, colWidths=[0.8*inch, 1*inch, 1.5*inch, 1.2*inch, 1*inch])
            trade_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27ae60')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(trade_table)
        
        elements.append(Spacer(1, 0.2*inch))

        # Seção 4: KYC
        if export_data['kyc']:
            kyc = export_data['kyc']
            elements.append(Paragraph("<b>4. VERIFICAÇÃO KYC</b>", styles['Heading2']))
            kyc_text = f"<b>Status:</b> {kyc['status']}<br/><b>Verificado em:</b> {kyc['verified_at']}"
            if kyc['daily_limit']:
                kyc_text += f"<br/><b>Limite Diário:</b> R$ {kyc['daily_limit']:.2f}"
            elements.append(Paragraph(kyc_text, styles['Normal']))
            elements.append(Spacer(1, 0.2*inch))

        # Rodapé
        elements.append(Spacer(1, 0.3*inch))
        footer_text = "Este relatório contém informações sensíveis. Mantenha-o seguro. Este documento foi gerado automaticamente."
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER,
        )
        elements.append(Paragraph(footer_text, footer_style))

        # Gerar PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def export_to_excel(user: User, export_data: Dict[str, Any]) -> bytes:
        """Exporta dados para Excel com múltiplas abas"""
        if not OPENPYXL_AVAILABLE:
            logger.error("❌ OpenPyXL não instalado")
            raise ImportError("OpenPyXL não instalado. Execute: pip install openpyxl")

        logger.info("📊 Gerando Excel...")

        wb = Workbook()
        wb.remove(wb.active)  # Remove sheet padrão

        # Estilos
        header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Sheet 1: Perfil
        ws_profile = wb.create_sheet("Perfil da Conta")
        profile = export_data['user_profile']
        profile_data = [
            ['Campo', 'Valor'],
            ['Username', profile['username']],
            ['Email', profile['email']],
            ['Criado em', profile['account_created']],
            ['Status', profile['account_status']],
            ['Email Verificado', 'Sim' if profile['email_verified'] else 'Não'],
            ['2FA Ativado', 'Sim' if profile['two_fa_enabled'] else 'Não'],
            ['Last Login', profile['last_login'] or 'Nunca'],
        ]
        
        for row in profile_data:
            ws_profile.append(row)
        
        # Formatar headers
        for cell in ws_profile[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border

        # Sheet 2: Carteiras
        ws_wallets = wb.create_sheet("Carteiras")
        wallet_headers = ['Crypto', 'Saldo', 'Saldo (BRL)', 'Endereço', 'Criada em', 'Status']
        ws_wallets.append(wallet_headers)
        
        for wallet in export_data['wallets']['wallets']:
            ws_wallets.append([
                wallet['crypto'],
                f"${wallet['balance']:.4f}",
                f"R$ {wallet['balance_in_brl']:.2f}",
                wallet['address'],
                wallet['created_at'],
                'Ativa' if wallet['is_active'] else 'Inativa',
            ])

        for cell in ws_wallets[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Sheet 3: Trades OTC
        ws_trades = wb.create_sheet("Trades OTC")
        trade_headers = ['Tipo', 'Status', 'Quantidade', 'Valor (BRL)', 'Crypto', 'Data', 'Concluída em']
        ws_trades.append(trade_headers)
        
        for trade in export_data['trades']['trades']:
            ws_trades.append([
                trade['type'].upper(),
                trade['status'],
                f"{trade['quantity']:.4f}",
                f"R$ {trade['amount']:.2f}",
                trade['crypto'],
                trade['created_at'],
                trade['completed_at'] or '',
            ])

        for cell in ws_trades[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Sheet 4: Pedidos P2P
        if export_data['p2p']['orders']:
            ws_p2p = wb.create_sheet("P2P")
            p2p_headers = ['Tipo', 'Status', 'Quantidade', 'Preço', 'Valor Total', 'Crypto', 'Data']
            ws_p2p.append(p2p_headers)
            
            for order in export_data['p2p']['orders']:
                ws_p2p.append([
                    order['order_type'],
                    order['status'],
                    f"{order['amount']:.4f}",
                    f"R$ {order['price']:.2f}",
                    f"R$ {order['total_value']:.2f}",
                    order['crypto'],
                    order['created_at'],
                ])

            for cell in ws_p2p[1]:
                cell.fill = header_fill
                cell.font = header_font

        # Sheet 5: KYC
        if export_data['kyc']:
            ws_kyc = wb.create_sheet("KYC")
            kyc = export_data['kyc']
            kyc_data = [
                ['Campo', 'Valor'],
                ['Status', kyc['status']],
                ['Documento', kyc['document_type']],
                ['Verificado em', kyc['verified_at'] or ''],
                ['Expira em', kyc['expires_at'] or ''],
                ['Limite Diário', f"R$ {kyc['daily_limit'] or 0:.2f}"],
                ['Limite Mensal', f"R$ {kyc['monthly_limit'] or 0:.2f}"],
            ]
            
            for row in kyc_data:
                ws_kyc.append(row)

            for cell in ws_kyc[1]:
                cell.fill = header_fill
                cell.font = header_font

        # Ajustar largura das colunas
        for ws in wb.sheetnames:
            sheet = wb[ws]
            for column in sheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
